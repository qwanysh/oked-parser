import time

import httpx
import openpyxl


def get_oked_by_bin(bin):
    url = f'https://stat.gov.kz/api/juridicalusr/counter/gov/?bin={bin}&lang=ru'
    headers = {
        'Accept': 'application/json',
        'Referer': 'https://stat.gov.kz/jur-search/bin',
    }

    while True:
        response = httpx.get(url, headers=headers)

        if response.status_code == httpx.codes.OK:
            response_json = response.json()
            if response_json['success']:
                return response_json['obj']['okedCode']
        elif response.status_code == httpx.codes.BAD_REQUEST:
            break

        time.sleep(0.2)


def get_column_index_by_value(worksheet, value):
    for index, cell in enumerate(next(worksheet.rows)):
        if cell.value == value:
            return index
    return -1


def main():
    workbook = openpyxl.load_workbook('companies.xlsx')
    worksheet = workbook.active
    bin_column_index = get_column_index_by_value(worksheet, value='bin')

    oked_column_index = get_column_index_by_value(worksheet, value='oked')
    if oked_column_index == -1:
        oked_column_index = worksheet.max_column
        worksheet.cell(row=1, column=oked_column_index + 1).value = 'oked'

    meta_column_index = get_column_index_by_value(worksheet, value='meta')
    if meta_column_index == -1:
        meta_column_index = oked_column_index + 1
        worksheet.cell(row=1, column=meta_column_index + 1).value = 'meta'

    workbook.save('companies.xlsx')

    for row, row_cells in enumerate(worksheet.rows, start=1):
        if row == 1 or row_cells[meta_column_index].value:
            continue

        bin = row_cells[bin_column_index].value
        oked = get_oked_by_bin(bin)
        worksheet.cell(row=row, column=oked_column_index + 1).value = oked or '-'
        worksheet.cell(row=row, column=meta_column_index + 1).value = 'parsed'
        workbook.save('companies.xlsx')


if __name__ == '__main__':
    start_time = time.time()
    main()
    print('total time:', time.time() - start_time)
